"""
excel_service.py — Generate templates, parse uploads, export current data for
TOU rate schedules and plant load profiles.

TOU Rate Excel structure:
  Sheet "TOU_Rates":   day_type | period_type | time_start | time_end | rate_baht_per_kwh | ft_baht_per_kwh
  Sheet "Demand_Charges": demand_charge_baht_per_kw_month | contract_demand_kw | service_fee_baht_per_month
  Sheet "Instructions": human-readable guide

Plant Load Excel structure:
  Sheet "Plant_Load":  minute (0-1439) | time (HH:MM) | load_kw
  Sheet "Spike_Events": start_tod | end_tod | extra_kw | probability
  Sheet "Instructions": human-readable guide
"""

from __future__ import annotations

import io
import json
from typing import Any

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ── Brand colours ──────────────────────────────────────────────────────────────
_RED   = "E3000F"
_DARK  = "18181B"
_GRAY  = "27272A"
_WHITE = "FAFAFA"
_MUTED = "A1A1AA"
_BORDER_COLOR = "3F3F46"

# ── Style helpers ──────────────────────────────────────────────────────────────

def _header_font() -> Font:
    return Font(bold=True, color=_WHITE, name="Calibri", size=10)

def _label_font() -> Font:
    return Font(color=_MUTED, name="Calibri", size=9, italic=True)

def _data_font() -> Font:
    return Font(color=_WHITE, name="Calibri", size=10)

def _header_fill() -> PatternFill:
    return PatternFill("solid", fgColor=_RED)

def _alt_fill() -> PatternFill:
    return PatternFill("solid", fgColor=_GRAY)

def _dark_fill() -> PatternFill:
    return PatternFill("solid", fgColor=_DARK)

def _thin_border() -> Border:
    s = Side(style="thin", color=_BORDER_COLOR)
    return Border(left=s, right=s, top=s, bottom=s)

def _apply_header_row(ws, row: int, columns: list[str]) -> None:
    for col_idx, title in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=col_idx, value=title)
        cell.font = _header_font()
        cell.fill = _header_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = max(18, len(title) + 4)

def _apply_data_row(ws, row: int, values: list[Any], alt: bool = False) -> None:
    fill = _alt_fill() if alt else _dark_fill()
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font = _data_font()
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _thin_border()

def _add_instruction_row(ws, row: int, text: str) -> None:
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(color=_MUTED, name="Calibri", size=9)
    cell.fill = _dark_fill()
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)

def _style_sheet(ws) -> None:
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = _RED


# ══════════════════════════════════════════════════════════════════════════════
# TOU RATE — Template & Export
# ══════════════════════════════════════════════════════════════════════════════

_TOU_COLS = ["day_type", "period_type", "time_start", "time_end",
             "rate_baht_per_kwh", "ft_baht_per_kwh"]
_DEMAND_COLS = ["demand_charge_baht_per_kw_month", "contract_demand_kw",
                "service_fee_baht_per_month"]

_TOU_TEMPLATE_ROWS = [
    ["weekday", "on_peak",  "09:00", "22:00", 4.1839, 0.0972],
    ["weekday", "off_peak", "00:00", "09:00", 2.6037, 0.0972],
    ["weekday", "off_peak", "22:00", "24:00", 2.6037, 0.0972],
    ["weekend", "off_peak", "00:00", "24:00", 2.6037, 0.0972],
]
_DEMAND_TEMPLATE_ROWS = [
    [132.93, 1600, 312.24],
]

_TOU_INSTRUCTIONS = [
    "INSTRUCTIONS — TOU Rate Schedule",
    "",
    "Sheet 'TOU_Rates':",
    "  • day_type     : 'weekday' or 'weekend' (holiday treated as weekend by optimizer)",
    "  • period_type  : 'on_peak' or 'off_peak'",
    "  • time_start   : HH:MM  (24-hour, e.g. 09:00)",
    "  • time_end     : HH:MM  (use 24:00 for midnight end-of-day)",
    "  • rate_baht_per_kwh : Energy charge base rate (Baht/kWh, EXCLUDING FT adder)",
    "  • ft_baht_per_kwh   : Fuel Tariff adder (Baht/kWh)",
    "",
    "Sheet 'Demand_Charges':",
    "  • demand_charge_baht_per_kw_month : Monthly demand charge (Baht/kW/month)",
    "  • contract_demand_kw              : Plant contract demand limit (kW)",
    "  • service_fee_baht_per_month      : Fixed monthly service fee (Baht/month)",
    "",
    "Upload behavior:",
    "  • Uploading replaces TOU settings in the database immediately.",
    "  • The optimizer reads the on_peak weekday row for peak_hours_start / peak_hours_end.",
    "  • On-peak rate = rate_baht_per_kwh (first on_peak weekday row found).",
    "  • Off-peak rate = rate_baht_per_kwh (first off_peak weekday row found).",
    "  • FT adder is taken from the first row (must be the same across all rows).",
]


def generate_tou_template() -> io.BytesIO:
    """Blank TOU rate template with default Sharp/MEA TOU 4.2.2 values."""
    return _build_tou_workbook(_TOU_TEMPLATE_ROWS, _DEMAND_TEMPLATE_ROWS)


def export_tou_to_excel(settings_dict: dict[str, str]) -> io.BytesIO:
    """Export current TOU settings from DB as filled Excel."""
    def get(key: str, default: str = "") -> str:
        return settings_dict.get(key, default)

    on_rate  = float(get("tou_onpeak_baht_per_kwh",  "4.1839"))
    off_rate = float(get("tou_offpeak_baht_per_kwh", "2.6037"))
    ft       = float(get("ft_baht_per_kwh",          "0.0972"))
    pk_start = get("peak_hours_start", "09:00")
    pk_end   = get("peak_hours_end",   "22:00")

    # Reconstruct rows from stored scalar values
    tou_rows = [
        ["weekday", "on_peak",  pk_start, pk_end,  on_rate,  ft],
        ["weekday", "off_peak", "00:00",  pk_start, off_rate, ft],
        ["weekday", "off_peak", pk_end,   "24:00",  off_rate, ft],
        ["weekend", "off_peak", "00:00",  "24:00",  off_rate, ft],
    ]
    demand_rows = [[
        float(get("demand_charge_baht_per_kw_month", "132.93")),
        float(get("contract_demand_kw",              "1600")),
        312.24,
    ]]

    return _build_tou_workbook(tou_rows, demand_rows)


def _build_tou_workbook(tou_rows: list, demand_rows: list) -> io.BytesIO:
    wb = openpyxl.Workbook()

    # ── Sheet 1: TOU_Rates ──────────────────────────────────────────────────
    ws_tou = wb.active
    ws_tou.title = "TOU_Rates"
    _style_sheet(ws_tou)
    ws_tou.row_dimensions[1].height = 22
    _apply_header_row(ws_tou, 1, _TOU_COLS)
    for i, row in enumerate(tou_rows):
        ws_tou.row_dimensions[i + 2].height = 18
        _apply_data_row(ws_tou, i + 2, row, alt=(i % 2 == 1))

    # ── Sheet 2: Demand_Charges ─────────────────────────────────────────────
    ws_dem = wb.create_sheet("Demand_Charges")
    _style_sheet(ws_dem)
    ws_dem.row_dimensions[1].height = 22
    _apply_header_row(ws_dem, 1, _DEMAND_COLS)
    for i, row in enumerate(demand_rows):
        ws_dem.row_dimensions[i + 2].height = 18
        _apply_data_row(ws_dem, i + 2, row, alt=False)
    ws_dem.column_dimensions["A"].width = 34
    ws_dem.column_dimensions["B"].width = 22
    ws_dem.column_dimensions["C"].width = 28

    # ── Sheet 3: Instructions ───────────────────────────────────────────────
    ws_ins = wb.create_sheet("Instructions")
    _style_sheet(ws_ins)
    ws_ins.column_dimensions["A"].width = 80
    for i, line in enumerate(_TOU_INSTRUCTIONS, start=1):
        ws_ins.row_dimensions[i].height = 15
        _add_instruction_row(ws_ins, i, line)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── TOU parse ──────────────────────────────────────────────────────────────────

def parse_tou_excel(file_bytes: bytes) -> dict[str, str]:
    """
    Parse TOU rate Excel upload.
    Returns dict keyed by settings config_key ready to upsert into the DB.
    Raises ValueError on invalid structure.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    if "TOU_Rates" not in wb.sheetnames:
        raise ValueError("Missing sheet 'TOU_Rates'")

    ws = wb["TOU_Rates"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    # Filter out fully empty rows
    rows = [r for r in rows if any(v is not None for v in r)]
    if not rows:
        raise ValueError("Sheet 'TOU_Rates' has no data rows")

    # Locate header columns by name (row 1)
    header_row = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
    def col(name: str) -> int:
        try:
            return header_row.index(name)
        except ValueError:
            raise ValueError(f"Missing column '{name}' in TOU_Rates sheet")

    i_day   = col("day_type")
    i_ptype = col("period_type")
    i_ts    = col("time_start")
    i_te    = col("time_end")
    i_rate  = col("rate_baht_per_kwh")
    i_ft    = col("ft_baht_per_kwh")

    on_peak_weekday  = None
    off_peak_weekday = None
    ft_val           = None

    for r in rows:
        def cell(idx: int) -> str:
            v = r[idx] if idx < len(r) else None
            return str(v).strip() if v is not None else ""

        day   = cell(i_day).lower()
        ptype = cell(i_ptype).lower()
        ts    = cell(i_ts)
        te    = cell(i_te)
        rate  = r[i_rate] if i_rate < len(r) else None
        ft    = r[i_ft]   if i_ft  < len(r) else None

        if ft is not None and ft_val is None:
            ft_val = float(ft)

        if day == "weekday" and ptype == "on_peak" and on_peak_weekday is None:
            on_peak_weekday = {"time_start": ts, "time_end": te, "rate": float(rate)}

        if day == "weekday" and ptype == "off_peak" and off_peak_weekday is None:
            off_peak_weekday = {"rate": float(rate)}

    if on_peak_weekday is None:
        raise ValueError("No 'weekday / on_peak' row found in TOU_Rates sheet")
    if off_peak_weekday is None:
        raise ValueError("No 'weekday / off_peak' row found in TOU_Rates sheet")

    result: dict[str, str] = {
        "tou_onpeak_baht_per_kwh":  str(on_peak_weekday["rate"]),
        "tou_offpeak_baht_per_kwh": str(off_peak_weekday["rate"]),
        "peak_hours_start":         on_peak_weekday["time_start"],
        "peak_hours_end":           on_peak_weekday["time_end"],
    }
    if ft_val is not None:
        result["ft_baht_per_kwh"] = str(ft_val)

    # Demand_Charges sheet (optional)
    if "Demand_Charges" in wb.sheetnames:
        ws_dem = wb["Demand_Charges"]
        dem_header = [str(c.value).strip().lower() if c.value else "" for c in ws_dem[1]]
        dem_rows = list(ws_dem.iter_rows(min_row=2, values_only=True))
        dem_rows = [r for r in dem_rows if any(v is not None for v in r)]
        if dem_rows:
            r0 = dem_rows[0]
            def dcol(name: str) -> int | None:
                try:
                    return dem_header.index(name)
                except ValueError:
                    return None
            di_dc = dcol("demand_charge_baht_per_kw_month")
            di_cd = dcol("contract_demand_kw")
            if di_dc is not None and r0[di_dc] is not None:
                result["demand_charge_baht_per_kw_month"] = str(float(r0[di_dc]))
            if di_cd is not None and r0[di_cd] is not None:
                result["contract_demand_kw"] = str(float(r0[di_cd]))

    return result


# ══════════════════════════════════════════════════════════════════════════════
# PLANT LOAD — Template & Export
# ══════════════════════════════════════════════════════════════════════════════

_PLANT_COLS  = ["minute", "time", "load_kw"]
_SPIKE_COLS  = ["start_tod", "end_tod", "extra_kw", "probability"]

_DEFAULT_STEP_PROFILE = [
    # (tod_start_min, tod_end_min, load_kw)
    (0,    7*60,  450.0),
    (7*60, 12*60, 800.0),
    (12*60,13*60, 600.0),
    (13*60,17*60, 850.0),
    (17*60,22*60, 750.0),
    (22*60,24*60, 450.0),
]

_DEFAULT_SPIKES = [
    ["10:30", "11:00", 400.0, 1.0],
    ["15:00", "15:30", 400.0, 1.0],
]

_PLANT_INSTRUCTIONS = [
    "INSTRUCTIONS — Plant Load Profile",
    "",
    "Sheet 'Plant_Load':",
    "  • minute   : Integer 0–1439 (minutes since midnight, 0 = 00:00)",
    "  • time     : HH:MM label (informational, not parsed — minute column is used)",
    "  • load_kw  : Baseline plant load at that minute (kW, excluding furnace load)",
    "  • All 1440 rows (minute 0–1439) should be present.",
    "    Missing rows default to the nearest preceding value.",
    "",
    "Sheet 'Spike_Events' (optional):",
    "  • start_tod   : HH:MM — spike window start (time-of-day)",
    "  • end_tod     : HH:MM — spike window end",
    "  • extra_kw    : Additional load during spike (kW)",
    "  • probability : 0.0–1.0 (1.0 = always occurs in simulation)",
    "",
    "Upload behavior:",
    "  • Uploading replaces the active plant load profile immediately.",
    "  • The optimizer uses minute-by-minute load_kw to compute demand headroom.",
    "  • Spike events are stored and optionally injected during simulation.",
]


def _default_minute_entries() -> list[dict]:
    entries = []
    for minute in range(1440):
        load_kw = 450.0
        for tod_start, tod_end, kw in _DEFAULT_STEP_PROFILE:
            if tod_start <= minute < tod_end:
                load_kw = kw
                break
        entries.append({"minute": minute, "load_kw": load_kw})
    return entries


def generate_plant_load_template() -> io.BytesIO:
    """Blank plant load template with default step-profile values (1440 rows)."""
    entries = _default_minute_entries()
    return _build_plant_load_workbook(entries, _DEFAULT_SPIKES)


def export_plant_load_to_excel(entries_json: str, spikes_json: str = "[]") -> io.BytesIO:
    """Export active plant load profile from DB as Excel."""
    entries = json.loads(entries_json) if entries_json else _default_minute_entries()
    spikes  = json.loads(spikes_json)  if spikes_json  else []
    spike_rows = [[s.get("start_tod",""), s.get("end_tod",""),
                   s.get("extra_kw", 0), s.get("probability", 1.0)] for s in spikes]
    return _build_plant_load_workbook(entries, spike_rows)


def _build_plant_load_workbook(entries: list[dict], spike_rows: list) -> io.BytesIO:
    wb = openpyxl.Workbook()

    # ── Sheet 1: Plant_Load ─────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Plant_Load"
    _style_sheet(ws)
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    _apply_header_row(ws, 1, _PLANT_COLS)
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 14

    for i, entry in enumerate(entries):
        minute = entry["minute"]
        hh, mm = divmod(minute, 60)
        time_str = f"{hh:02d}:{mm:02d}"
        ws.row_dimensions[i + 2].height = 15
        _apply_data_row(ws, i + 2, [minute, time_str, entry["load_kw"]], alt=(i % 2 == 1))

    # ── Sheet 2: Spike_Events ───────────────────────────────────────────────
    ws_spk = wb.create_sheet("Spike_Events")
    _style_sheet(ws_spk)
    ws_spk.row_dimensions[1].height = 22
    _apply_header_row(ws_spk, 1, _SPIKE_COLS)
    for col_letter, width in zip(["A","B","C","D"], [14, 14, 14, 14]):
        ws_spk.column_dimensions[col_letter].width = width
    for i, row in enumerate(spike_rows):
        ws_spk.row_dimensions[i + 2].height = 18
        _apply_data_row(ws_spk, i + 2, row, alt=(i % 2 == 1))

    # ── Sheet 3: Instructions ───────────────────────────────────────────────
    ws_ins = wb.create_sheet("Instructions")
    _style_sheet(ws_ins)
    ws_ins.column_dimensions["A"].width = 80
    for i, line in enumerate(_PLANT_INSTRUCTIONS, start=1):
        ws_ins.row_dimensions[i].height = 15
        _add_instruction_row(ws_ins, i, line)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Plant load parse ───────────────────────────────────────────────────────────

def parse_plant_load_excel(file_bytes: bytes) -> tuple[list[dict], list[dict]]:
    """
    Parse plant load Excel upload.
    Returns (entries, spikes) where:
      entries = [{minute: int, load_kw: float}, ...]  (sorted 0..1439)
      spikes  = [{start_tod, end_tod, extra_kw, probability}, ...]
    Raises ValueError on invalid structure.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    if "Plant_Load" not in wb.sheetnames:
        raise ValueError("Missing sheet 'Plant_Load'")

    ws = wb["Plant_Load"]
    header = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]

    def col(name: str) -> int:
        try:
            return header.index(name)
        except ValueError:
            raise ValueError(f"Missing column '{name}' in Plant_Load sheet")

    i_min = col("minute")
    i_kw  = col("load_kw")

    raw: dict[int, float] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        m_val  = row[i_min] if i_min < len(row) else None
        kw_val = row[i_kw]  if i_kw  < len(row) else None
        if m_val is None or kw_val is None:
            continue
        minute = int(m_val)
        if not (0 <= minute <= 1439):
            continue
        raw[minute] = float(kw_val)

    if len(raw) == 0:
        raise ValueError("No valid data rows found in Plant_Load sheet")

    # Fill any missing minutes by forward-fill from nearest preceding value
    entries: list[dict] = []
    last_kw = raw.get(0, 450.0)
    for minute in range(1440):
        if minute in raw:
            last_kw = raw[minute]
        entries.append({"minute": minute, "load_kw": last_kw})

    # Spikes sheet (optional)
    spikes: list[dict] = []
    if "Spike_Events" in wb.sheetnames:
        ws_spk = wb["Spike_Events"]
        spk_header = [str(c.value).strip().lower() if c.value else "" for c in ws_spk[1]]
        def scol(name: str) -> int | None:
            try:
                return spk_header.index(name)
            except ValueError:
                return None
        si_start = scol("start_tod")
        si_end   = scol("end_tod")
        si_kw    = scol("extra_kw")
        si_prob  = scol("probability")
        for row in ws_spk.iter_rows(min_row=2, values_only=True):
            if all(v is None for v in row):
                continue
            spike: dict = {}
            if si_start is not None and row[si_start] is not None:
                spike["start_tod"] = str(row[si_start]).strip()
            if si_end is not None and row[si_end] is not None:
                spike["end_tod"] = str(row[si_end]).strip()
            if si_kw is not None and row[si_kw] is not None:
                spike["extra_kw"] = float(row[si_kw])
            if si_prob is not None and row[si_prob] is not None:
                spike["probability"] = float(row[si_prob])
            if spike:
                spikes.append(spike)

    return entries, spikes
